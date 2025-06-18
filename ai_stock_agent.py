# ai_stock_agent.py

import yfinance as yf
import time
from datetime import datetime
import winsound
import csv
import os
from rich.console import Console
from rich.table import Table
from rich.live import Live
from rich import box
import requests
from bs4 import BeautifulSoup
from nltk.sentiment.vader import SentimentIntensityAnalyzer
import nltk
from plyer import notification
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

nltk.download('vader_lexicon')

console = Console()

class StockAIAgent:
    def __init__(self, symbols, refresh_interval=10):
        self.symbols = symbols
        self.refresh_interval = refresh_interval
        self.previous_prices = {}
        self.alerted_symbols = set()
        self.sentiment_analyzer = SentimentIntensityAnalyzer()
        self.PRICE_THRESHOLDS = {
            "AAPL": {"min": 100, "max": 150},
            "TSLA": {"min": 150, "max": 350},
            "GOOGL": {"min": 100, "max": 200},
        }
        self.CSV_FILE = "stock-market-prices.csv"
        self.EXCEL_FILE = "stock_prices.xlsx"
        self.daily_summary = []

    def init_csv(self):
        if not os.path.exists(self.CSV_FILE):
            with open(self.CSV_FILE, mode="w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Timestamp", "Symbol", "Name", "Price"])

    def export_to_csv(self, timestamp, rows):
        with open(self.CSV_FILE, mode="a", newline="") as f:
            writer = csv.writer(f)
            for row in rows:
                writer.writerow([timestamp] + row)

    def export_to_excel(self, timestamp, rows):
        if not os.path.exists(self.EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Prices"
            ws.append(["Timestamp", "Symbol", "Name", "Price"])
        else:
            wb = load_workbook(self.EXCEL_FILE)
            ws = wb["Prices"]
        for row in rows:
            ws.append([timestamp] + row)
        wb.save(self.EXCEL_FILE)

    def add_chart_to_excel(self):
        try:
            wb = load_workbook(self.EXCEL_FILE)
            ws = wb["Prices"]
            chart = LineChart()
            chart.title = "Stock Price Trend"
            chart.y_axis.title = "Price"
            chart.x_axis.title = "Time"
            data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, f"G2")
            wb.save(self.EXCEL_FILE)
        except Exception as e:
            console.print(f"Chart Error: {e}")

    def fetch_sentiment(self, symbol):
        try:
            url = f"https://finance.yahoo.com/quote/{symbol}?p={symbol}"
            response = requests.get(url, timeout=5)
            soup = BeautifulSoup(response.text, "html.parser")
            headline_tag = soup.find("li", {"data-test-locator": "mega"})
            headline = headline_tag.text.strip() if headline_tag else "No headline"
            sentiment = self.sentiment_analyzer.polarity_scores(headline)['compound']
            return headline, sentiment
        except Exception:
            return "News error", 0.0

    def classify_action(self, price, sentiment, thresholds):
        if sentiment > 0.5 and price < thresholds.get("max", float("inf")):
            return "🟢 Buy"
        elif sentiment < -0.5 and price > thresholds.get("min", float("-inf")):
            return "🔴 Sell"
        else:
            return "🟡 Hold"

    def send_notification(self, title, message):
        notification.notify(title=title, message=message, timeout=5)
        try:
            winsound.PlaySound("sound.mp3", winsound.SND_ALIAS)
        except Exception as e:
            print(f"Sound Error: {e}")

    def build_table(self):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        table = Table(title=f"🤖 AI Stock Agent | {timestamp}", box=box.SQUARE)
        table.add_column("Symbol", style="bold")
        table.add_column("Name")
        table.add_column("Price", justify="right")
        table.add_column("Action", justify="center")
        table.add_column("Alert", justify="left")
        table.add_column("News", justify="left", overflow="fold")

        csv_rows = []
        self.daily_summary.clear()

        for symbol in self.symbols:
            try:
                stock = yf.Ticker(symbol)
                info = stock.info
                price = info['regularMarketPrice']
                name = info.get('shortName', 'Unknown')

                prev_price = self.previous_prices.get(symbol)
                self.previous_prices[symbol] = price

                headline, sentiment = self.fetch_sentiment(symbol)
                mood = "🟢" if sentiment >= 0.2 else "🔴" if sentiment <= -0.2 else "🟡"

                thresholds = self.PRICE_THRESHOLDS.get(symbol, {})
                action = self.classify_action(price, sentiment, thresholds)

                alert = ""
                if thresholds:
                    if price > thresholds.get("max", float("inf")) and symbol not in self.alerted_symbols:
                        alert = "🔺 ABOVE MAX"
                        self.send_notification(f"{symbol} Price Alert", f"{symbol} is ABOVE MAX: ${price}")
                        self.alerted_symbols.add(symbol)
                    elif price < thresholds.get("min", float("-inf")) and symbol not in self.alerted_symbols:
                        alert = "🔻 BELOW MIN"
                        self.send_notification(f"{symbol} Price Alert", f"{symbol} is BELOW MIN: ${price}")
                        self.alerted_symbols.add(symbol)
                    elif thresholds.get("min") < price < thresholds.get("max"):
                        if symbol in self.alerted_symbols:
                            self.alerted_symbols.remove(symbol)

                table.add_row(symbol, name, f"${price:.2f}", action, alert, f"{mood} {headline[:60]}...")
                csv_rows.append([symbol, name, price])
                self.daily_summary.append(f"{symbol}: {name}, Price=${price:.2f}, Action={action}, Sentiment={mood}")

            except Exception as e:
                table.add_row(symbol, "⚠️ Error", "N/A", "–", "", str(e))

        self.export_to_csv(timestamp, csv_rows)
        self.export_to_excel(timestamp, csv_rows)
        return table

    def send_daily_summary_email(self, recipient_email):
        sender_email = "alishermaan0319@gmail.com"  # Replace with your email
        sender_password = "sherjatt2319"
        subject = "📈 Daily Stock Summary from Your AI Agent"
        body = "\n".join(self.daily_summary)

        message = MIMEMultipart()
        message['From'] = sender_email
        message['To'] = recipient_email
        message['Subject'] = subject
        message.attach(MIMEText(body, 'plain'))

        try:
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(message)
            server.quit()
            console.print("✅ Daily summary email sent.", style="bold green")
        except Exception as e:
            console.print(f"📧 Email Error: {e}", style="bold red")

    def run(self):
        self.init_csv()
        with Live(self.build_table(), refresh_per_second=1, screen=True) as live:
            try:
                while True:
                    time.sleep(self.refresh_interval)
                    live.update(self.build_table())
            except KeyboardInterrupt:
                console.print("\n⛔ Exiting AI Agent...", style="bold red")
                self.add_chart_to_excel()
                self.send_daily_summary_email("alishermaan0319@gmail.com")  # Replace with your email

if __name__ == "__main__":
    user_input = input("Enter stock symbols separated by commas (or press Enter to use defaults): ").strip()
    symbols = [s.strip().upper() for s in user_input.split(",")] if user_input else [
        "AAPL", "TSLA", "AMZN", "GOOGL", "MSFT", "NVDA", "META",
        "NFLX", "IBM", "INTC", "BABA", "BA", "JPM", "DIS", "ORCL",
        "PYPL", "ADBE", "PEP", "NKE", "KO"]

    try:
        refresh = int(input("Enter refresh interval in seconds (e.g., 10): "))
    except ValueError:
        refresh = 10
        print("⚠️ Invalid input. Using default interval: 10 seconds.")

    agent = StockAIAgent(symbols, refresh)
    agent.run()
# Ensure you have the required packages installed:
# pip install yfinance rich nltk plyer openpyxl requests beautifulsoup4 smtplib email   
# Note: Replace 
